from openpyxl import load_workbook

def read_table(path):
    try:
        wb = load_workbook(path)
        sheet = wb.active
    except:
        raise Exception("бля я чет не могу открыть эту хуйню, ты допустил ошибку, проверся на спид")

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data.append(row)

    return data
